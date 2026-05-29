#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Analyze ZYY字段名与属性.xlsx and compare with 输出结果.gdb
"""

import os
import sys

# ---- Step 1: Read xlsx ----
xlsx_path = r'c:\4code\3lot\ZYY字段名与属性.xlsx'
gdb_path = r'c:\4code\3lot\输出结果.gdb'

print("=" * 80)
print("STEP 1: Reading xlsx file")
print("=" * 80)

# Try openpyxl first, then fallback
try:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Using openpyxl to read: {xlsx_path}")
    for ws in wb.worksheets:
        print(f"\n--- Sheet: {ws.title} (rows={ws.max_row}, cols={ws.max_column}) ---")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            vals = [cell.value for cell in row]
            print(f"  {vals}")
    wb.close()
except ImportError:
    print("openpyxl not available, trying pandas...")
    try:
        import pandas as pd
        xls = pd.ExcelFile(xlsx_path)
        for sheet_name in xls.sheet_names:
            print(f"\n--- Sheet: {sheet_name} ---")
            df = pd.read_excel(xls, sheet_name=sheet_name)
            print(df.to_string())
    except ImportError:
        print("pandas not available either, trying raw zip+xml...")
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            # shared strings
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                tree = ET.parse(z.open('xl/sharedStrings.xml'))
                root = tree.getroot()
                for si in root.findall('.//s:si', ns):
                    texts = [t.text or '' for t in si.findall('.//s:t', ns)]
                    shared_strings.append(''.join(texts))
                print(f"Shared strings: {len(shared_strings)}")

            # sheets
            for name in sorted(z.namelist()):
                if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                    print(f"\n--- {name} ---")
                    tree = ET.parse(z.open(name))
                    root = tree.getroot()
                    for row in root.findall('.//s:sheetData/s:row', ns):
                        row_num = row.get('r', '')
                        cells = []
                        for c in row.findall('s:c', ns):
                            ref = c.get('r', '')
                            ctype = c.get('t', '')
                            v = c.find('s:v', ns)
                            val = v.text if v is not None else ''
                            if ctype == 's' and val:
                                try:
                                    val = shared_strings[int(val)]
                                except (ValueError, IndexError):
                                    pass
                            cells.append(f"{ref}={val}")
                        print(f"  Row {row_num}: {' | '.join(cells)}")

if __name__ == '__main__':
    pass  # will run step by step below
