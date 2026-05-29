# -*- coding: utf-8 -*-
"""
可研附表模板自动填充脚本
从各县 ZZY.shp 读取林地图斑数据，按字段标准翻译代码，
填充到对应县的 可研附表模板.xlsx 的 A.1、B.1-B.6、Fee sheet 中。

用法: python populate_template.py
"""

import os, re, struct, warnings, zipfile
from copy import copy
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from project_config import OUTPUT_BASE, STANDARD_FILE, COUNTY_CODE_TO_NAME

warnings.filterwarnings("ignore")

BASE_DIR = OUTPUT_BASE
MD_FILE = STANDARD_FILE

SHEET_A1 = "表A.1建设项目使用林地因子调查表"
SHEET_B1 = "表B.1项目使用林地按林地类型面积蓄积统计表"
SHEET_B2 = "表B.2项目使用林地按地类面积统计表"
SHEET_B3 = "表B.3项目使用林地按森林类别面积统计表"
SHEET_B4 = "表B.4项目使用林地按林地保护等级面积统计表"
SHEET_B5 = "表B.5项目使用林地分森林类别按地类面积统计表"
SHEET_B6 = "表B.6项目使用重点生态区域林地面积统计表"
SHEET_FEE = "项目拟使用林地应缴纳森林植被恢复费测算统计表"

# ---- 保护区字段（仅检查ZRBHQ_MC） ----
PROTECTED_FIELDS = ["ZRBHQ_MC", "ZRBHQ_DJ"]

# ============================================================
# 工具函数
# ============================================================
def find_col(df, *kw):
    for c in df.columns:
        if all(k in str(c) for k in kw):
            return c
    return None

def _text(v):
    if v is None or pd.isna(v):
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null", "<null>") else s

def _norm_code(v, width=None):
    s = _text(v)
    if not s:
        return ""
    try:
        if re.fullmatch(r"-?\d+(\.0+)?", s):
            s = str(int(float(s)))
    except Exception:
        pass
    return s.zfill(width) if width and s.isdigit() else s

def _num(v):
    s = _text(v)
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def _has_value(v):
    return bool(_text(v))

def _dbf_encodings(dbf_path):
    encodings = []
    cpg = Path(dbf_path).with_suffix(".cpg")
    if cpg.exists():
        raw = cpg.read_text(errors="ignore").strip().lower()
        encodings.append({"utf8": "utf-8", "utf-8": "utf-8", "65001": "utf-8", "936": "gbk"}.get(raw, raw))
    encodings.extend(["utf-8", "gbk", "cp936"])
    out = []
    for enc in encodings:
        if enc and enc not in out:
            out.append(enc)
    return out

def read_dbf_table(path):
    """读取 shapefile 的 DBF 属性表；本脚本只统计属性，不需要几何。"""
    dbf_path = Path(path)
    if dbf_path.suffix.lower() != ".dbf":
        dbf_path = dbf_path.with_suffix(".dbf")
    if not dbf_path.exists():
        return pd.DataFrame()

    last_error = None
    for enc in _dbf_encodings(dbf_path):
        try:
            return _read_dbf_table_with_encoding(dbf_path, enc)
        except UnicodeDecodeError as e:
            last_error = e
    if last_error:
        raise last_error
    return pd.DataFrame()

def _read_dbf_table_with_encoding(dbf_path, encoding):
    data = Path(dbf_path).read_bytes()
    if len(data) < 32:
        return pd.DataFrame()
    record_count = struct.unpack("<I", data[4:8])[0]
    header_len = struct.unpack("<H", data[8:10])[0]
    record_len = struct.unpack("<H", data[10:12])[0]

    fields = []
    offset = 32
    while offset < header_len and data[offset] != 0x0D:
        desc = data[offset:offset + 32]
        name_raw = desc[:11].split(b"\x00", 1)[0]
        name = name_raw.decode(encoding).strip()
        ftype = chr(desc[11])
        flen = desc[16]
        fields.append((name, ftype, flen))
        offset += 32

    rows = []
    for i in range(record_count):
        start = header_len + i * record_len
        rec = data[start:start + record_len]
        if not rec or rec[:1] == b"*":
            continue
        pos = 1
        row = {}
        for name, _ftype, flen in fields:
            raw = rec[pos:pos + flen]
            pos += flen
            row[name] = raw.decode(encoding).replace("\x00", "").strip()
        rows.append(row)
    return pd.DataFrame(rows)

def _label_without_tree_marks(s):
    return re.sub(r"[├─│└┌┬┴┼┐┘┤ ]", "", _text(s)).strip()

def _subtotal_label(name):
    s = _text(name)
    if not s:
        return "小计"
    return s if s.endswith("小计") or s == "合计" else f"{s}小计"

def _normalize_ownership(v):
    s = _text(v)
    return "国有" if s in ("10", "国有") else "集体"

def _normalize_land_use(v):
    s = _text(v)
    if not s:
        return "其他林地"
    code = _norm_code(s, 3)
    if code.isdigit():
        n = int(code)
        if 111 <= n <= 117:
            return "防护林林地"
        if 121 <= n <= 127:
            return "特用林林地"
        if 231 <= n <= 233:
            return "用材林林地"
        if n == 240:
            return "能源林林地"
        if 251 <= n <= 255:
            return "经济林林地"

    lu_map = {
        "水源涵养林": "防护林林地", "水土保持林": "防护林林地", "防风固沙林": "防护林林地",
        "农田牧场防护林": "防护林林地", "护岸林": "防护林林地", "护路林": "防护林林地", "其他防护林": "防护林林地",
        "国防林": "特用林林地", "实验林": "特用林林地", "母树林": "特用林林地",
        "环境保护林": "特用林林地", "风景林": "特用林林地", "名胜古迹和革命纪念林": "特用林林地", "自然保护林": "特用林林地",
        "短轮伐期工业原料用材林": "用材林林地", "速生丰产用材林": "用材林林地", "一般用材林": "用材林林地",
        "果树林": "经济林林地", "食用原料林": "经济林林地", "林化工业原料林": "经济林林地",
        "药用林": "经济林林地", "其他经济林": "经济林林地", "能源林": "能源林林地",
    }
    if s in lu_map:
        return lu_map[s]
    for pre, res in [
        ("防护林", "防护林林地"), ("特种用途林", "特用林林地"),
        ("用材林", "用材林林地"), ("经济林", "经济林林地"), ("能源林", "能源林林地"),
    ]:
        if s.startswith(pre):
            return res
    return "其他林地"

def _is_blank_lin_zhong(row):
    return not _has_value(row.get("_raw_LIN_ZHONG", row.get("LIN_ZHONG", "")))

def _normalize_di_lei(v):
    s = _text(v)
    code = _norm_code(s, 6)
    code_map = {
        "030100": "乔木林地", "030101": "乔木林地",
        "030200": "竹林地",
        "030301": "特殊灌林地",
        "030302": "一般灌木林地", "030303": "一般灌木林地",
        "030401": "疏林地",
        "030402": "未成林地", "030403": "未成林地",
        "030404": "苗圃地",
        "030405": "采伐迹地",
        "030406": "火烧迹地",
    }
    if code in code_map:
        return code_map[code]
    if "乔木" in s:
        return "乔木林地"
    if "竹林" in s:
        return "竹林地"
    if "国家特别规定" in s or "特殊灌" in s:
        return "特殊灌林地"
    if "一般灌木" in s or "园地转灌木" in s:
        return "一般灌木林地"
    if "疏林" in s:
        return "疏林地"
    if "未成林" in s:
        return "未成林地"
    if "苗圃" in s:
        return "苗圃地"
    if "采伐" in s:
        return "采伐迹地"
    if "火烧" in s:
        return "火烧迹地"
    if "宜林" in s:
        return "宜林地"
    return "其他林地"

def _display_di_lei(v):
    n = _normalize_di_lei(v)
    return "特殊灌木林地" if n == "特殊灌林地" else n

def _display_origin(raw, translated=""):
    code = _norm_code(raw, 2)
    if code.startswith("1"):
        return "天然"
    if code.startswith("2"):
        return "人工"
    s = _text(translated) or _text(raw)
    if "天然" in s:
        return "天然"
    if "人工" in s or s in ("植苗", "直播", "飞播"):
        return "人工"
    return s

def _display_lingzu(v, is_bamboo=False):
    s = _text(v)
    if not s:
        return ""
    m = re.match(r"(.+?)[（(](.+?)[）)]", s)
    if not m:
        return s
    return m.group(2).strip() if is_bamboo else m.group(1).strip()

def _b5_di_lei(v):
    n = _normalize_di_lei(v)
    if n in ("乔木林地", "竹林地"):
        return n
    if n == "特殊灌林地":
        return "特殊灌木林地"
    return "其他林地"

def _normalize_protect_grade(v):
    s = _text(v).replace("保护", "")
    code = _norm_code(s)
    code_map = {"1": "Ⅰ级", "2": "Ⅱ级", "3": "Ⅲ级", "4": "Ⅳ级"}
    if code in code_map:
        return code_map[code]
    for g in code_map.values():
        if g in s:
            return g
    return s

def _display_protect_grade(v):
    g = _normalize_protect_grade(v)
    return f"{g}保护" if g else ""

def _normalize_sen(v):
    s = _text(v).replace("（地）", "地").replace("(地)", "地")
    code = _norm_code(s, 3)
    if code == "011":
        return "重点公益林地"
    if code == "012":
        return "一般公益林地"
    if code == "021":
        return "重点商品林地"
    if code == "022":
        return "一般商品林地"
    if "重点商品" in s:
        return "重点商品林地"
    if "一般商品" in s:
        return "一般商品林地"
    if "商品" in s:
        return "商品林地"
    if "公益" in s:
        return "公益林地"
    return s

def _normalize_sqdj(v):
    s = _text(v)
    code = _norm_code(s, 2)
    if code == "11" or "国家级" in s:
        return "国家级"
    if code == "21" or "省级" in s:
        return "省级"
    if code == "23" or "市级" in s:
        return "市级"
    if code == "24" or "县级" in s:
        return "县级"
    return ""

def _normalize_gjgy_level(v):
    s = _text(v)
    code = _norm_code(s)
    if code == "1" or "一级" in s:
        return "一级"
    if code == "2" or "二级" in s:
        return "二级"
    return ""

def _forest_leaf(sen, sqdj, gj_level):
    sq = _normalize_sqdj(sqdj)
    sen_n = _normalize_sen(sen)
    if sq == "国家级":
        gj = _normalize_gjgy_level(gj_level)
        return f"国家级{gj}" if gj else "国家级"
    if sq == "省级":
        return "省级"
    if sq in ("市级", "县级"):
        return "市县级"
    if "公益" in sen_n:
        return "市县级"
    if sen_n == "重点商品林地":
        return "重点商品"
    if sen_n in ("一般商品林地", "商品林地"):
        return "一般商品"
    return ""

def _forest_display(sen, sqdj, gj_level):
    leaf = _forest_leaf(sen, sqdj, gj_level)
    if leaf == "国家级一级":
        return "国家级一级公益林地"
    if leaf == "国家级二级":
        return "国家级二级公益林地"
    if leaf == "国家级":
        return "国家级公益林地"
    if leaf == "省级":
        return "省级公益林地"
    if leaf == "市县级":
        sq = _normalize_sqdj(sqdj)
        return f"{sq}公益林地" if sq in ("市级", "县级") else "其他公益林地"
    if leaf == "重点商品":
        return "重点商品林地"
    if leaf == "一般商品":
        return "一般商品林地"
    return _normalize_sen(sen)

def _b5_forest_row(leaf):
    if leaf.startswith("国家级"):
        return "国家级公益林地"
    if leaf == "省级":
        return "省级公益林地"
    if leaf == "市县级":
        return "其他公益林地"
    if leaf == "重点商品":
        return "重点商品林地"
    if leaf == "一般商品":
        return "一般商品林地"
    return ""

def _fee_public_class(leaf):
    return leaf.startswith("国家级") or leaf == "省级"

def _is_city_planning(v):
    return _text(v) in ("1", "是", "true", "True", "TRUE")

# ============================================================
# MD 字典解析（含乡镇/村名称表）
# ============================================================
def parse_md_dictionaries(md_path):
    with open(md_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    sections, cur, cur_lines = {}, None, []
    for line in lines:
        s = line.strip()
        if s.startswith("### "):
            if cur: sections[cur] = cur_lines
            cur, cur_lines = s, []
        elif cur is not None:
            cur_lines.append(s)
    if cur: sections[cur] = cur_lines

    name_map = {
        "一张图地类": "DI_LEI", "地类编码": "DLBM",
        "土地所有权属": "LD_QS", "林地保护等级": "BH_DJ",
        "森林类别": "SEN_LIN_LB", "林种": "LIN_ZHONG",
        "起源": "QI_YUAN", "优势树": "YOU_SHI_SZ",
        "龄组": "LING_ZU", "国家级公益林保护等级": "GJGYL_BHDJ",
        "事权等级": "Y_SQDJ", "油茶生产期": "YC_CQ",
    }
    dicts = {}
    for hdr, sl in sections.items():
        dn = None
        for kw, n in name_map.items():
            if kw in hdr: dn = n; break
        if dn is None: continue
        in_t = False
        rows = []
        for line in sl:
            if line.startswith("|") and line.endswith("|"):
                if not in_t: in_t = True; continue
                rows.append(line)
        cm = {}
        for row in rows:
            parts = [p.strip() for p in row.split("|") if p.strip()]
            if len(parts) >= 2:
                name = _label_without_tree_marks(parts[0])
                code = parts[1]
                if code and name: cm[code] = name
        if cm: dicts[dn] = cm
    return dicts

def _md_table_lines(content, title):
    m = re.search(rf"### {title}.*?(?=\n---|\n### |\Z)", content, re.DOTALL)
    if not m:
        return []
    return [line.strip() for line in m.group(0).splitlines()
            if line.strip().startswith("|") and line.strip().endswith("|")]

def parse_township_names(md_path):
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()
    names = {}
    for line in _md_table_lines(content, "乡镇代码"):
        parts = [p.strip() for p in line.split('|') if p.strip()]
        if len(parts) >= 4 and parts[1] != "县代码":
            names[(_norm_code(parts[1], 6), _norm_code(parts[2], 3))] = parts[3]
    return names

def parse_village_names(md_path):
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()
    names = {}
    for line in _md_table_lines(content, "行政村代码"):
        parts = [p.strip() for p in line.split('|') if p.strip()]
        if len(parts) >= 4 and parts[1] != "乡镇代码":
            county_name = parts[0]
            town_code = _norm_code(parts[1], 3)
            village_code = _norm_code(parts[2], 3)
            names[(county_name, town_code, village_code)] = parts[3]
            names[(town_code, village_code)] = parts[3]
    return names

# ============================================================
# 读 SHP
# ============================================================
def read_zzy_shp(path):
    if not os.path.exists(path): return None
    try:
        return read_dbf_table(path)
    except Exception as e:
        print(f"  读shp属性失败: {e}"); return None

# ============================================================
# _write_value — 安全写入，跳过0值
# ============================================================
def _write_value(ws, row, col, val, allow_zero=False):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return
    if not allow_zero and isinstance(val, (int, float, np.integer, np.floating)) and val == 0:
        return
    if isinstance(val, np.integer): val = int(val)
    elif isinstance(val, (float, np.floating)): val = round(float(val), 4)
    for _ in range(2):
        try:
            ws.cell(row=row, column=col).value = val
            return
        except AttributeError:
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    try: ws.unmerge_cells(str(mr)); break
                    except: pass
    try: ws.cell(row=row, column=col).value = val
    except: pass

def clear_data_area(ws, start_row):
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= start_row:
            try: ws.unmerge_cells(str(mr))
            except: pass
    for r in range(start_row, ws.max_row + 10):
        for c in range(1, ws.max_column + 1):
            try: ws.cell(row=r, column=c).value = None
            except: pass

def _copy_row_format(ws, src_row, dst_row):
    if src_row == dst_row:
        return
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)

def _clear_values(ws, start_row, end_row=None, start_col=1, end_col=None):
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            try:
                ws.cell(r, c).value = None
            except Exception:
                pass

def _prepare_dynamic_rows(ws, start_row, row_count, pattern_rows=3):
    end_row = start_row + max(row_count, 0) - 1
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= start_row:
            try:
                ws.unmerge_cells(str(mr))
            except Exception:
                pass
    if row_count > 0:
        for r in range(start_row, end_row + 1):
            src_row = start_row + ((r - start_row) % pattern_rows)
            if src_row <= ws.max_row:
                _copy_row_format(ws, src_row, r)
    _clear_values(ws, start_row, max(ws.max_row, end_row))

def _prepare_plain_rows(ws, start_row, row_count, pattern_row=None, start_col=1, end_col=None):
    pattern_row = pattern_row or start_row
    end_row = start_row + max(row_count, 0) - 1
    if row_count > 0:
        for r in range(start_row, end_row + 1):
            if pattern_row <= ws.max_row:
                _copy_row_format(ws, pattern_row, r)
    _clear_values(ws, start_row, max(ws.max_row, end_row), start_col=start_col, end_col=end_col)

def _merge_triplet_labels(ws, start_row, row_count):
    for r in range(start_row, start_row + row_count, 3):
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=1)
        except Exception:
            pass

def _reset_merges(ws, ranges):
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 6:
            try:
                ws.unmerge_cells(str(mr))
            except Exception:
                pass
    for rg in ranges:
        try:
            ws.merge_cells(rg)
        except Exception:
            pass

# ============================================================
# 翻译 ZZY → A.1
# ============================================================
def prepare_a1_data(zzy_gdf, dicts, c2n, town_names=None, vill_names=None):
    df = zzy_gdf.copy().reset_index(drop=True)
    for c in df.columns:
        if df[c].dtype == object: df[c] = df[c].fillna("").astype(str)

    # 字典翻译
    fd_map = {"DI_LEI":"DI_LEI","LD_QS":"LD_QS","BH_DJ":"BH_DJ","SEN_LIN_LB":"SEN_LIN_LB",
              "LIN_ZHONG":"LIN_ZHONG","QI_YUAN":"QI_YUAN","YOU_SHI_SZ":"YOU_SHI_SZ",
              "LING_ZU":"LING_ZU","Y_SQDJ":"Y_SQDJ","GJGYL_BHDJ":"GJGYL_BHDJ"}
    for f, dn in fd_map.items():
        if f in df.columns and dn in dicts:
            df[f"_raw_{f}"] = df[f]
            d = dicts[dn]
            def _m(x, lk=d):
                if isinstance(x, str):
                    s = x.strip()
                    return lk.get(s, x) if s in lk else x
                return x
            df[f] = df[f].map(_m)

    # ---- 县名 ----
    def _county(xc):
        code = _norm_code(xc, 6)
        return c2n.get(code, code)
    df["_xian_name"] = df["XIAN"].apply(_county) if "XIAN" in df.columns else ""

    # ---- 乡镇/村名 ----
    def _town(r):
        xg = _norm_code(r.get("XIANG", ""), 3)
        if town_names is None:
            return xg
        xc = _norm_code(r.get("XIAN", ""), 6)
        return town_names.get((xc, xg)) or xg
    def _vill(r):
        xg = _norm_code(r.get("XIANG", ""), 3)
        cn = _norm_code(r.get("CUN", ""), 3)
        if vill_names is None:
            return cn
        county_name = r.get("_xian_name", "")
        return vill_names.get((county_name, xg, cn)) or vill_names.get((xg, cn)) or cn
    df["_xiang_name"] = df.apply(_town, axis=1) if "XIANG" in df.columns else ""
    df["_cun_name"] = df.apply(_vill, axis=1) if "CUN" in df.columns else ""

    # ---- 林地权属 ----
    df["_ownership"] = df["LD_QS"].apply(_normalize_ownership) if "LD_QS" in df.columns else "集体"

    # ---- 使用林地类型 ----
    if "LIN_ZHONG" in df.columns:
        df["_land_use"] = df.apply(lambda r: _normalize_land_use(r.get("_raw_LIN_ZHONG", r.get("LIN_ZHONG", ""))) or _normalize_land_use(r.get("LIN_ZHONG", "")), axis=1)
    else:
        df["_land_use"] = "其他林地"

    # ---- 统计口径派生字段 ----
    if "DI_LEI" in df.columns:
        df["_di_lei_norm"] = df.apply(lambda r: _normalize_di_lei(r.get("_raw_DI_LEI", r.get("DI_LEI", ""))) or _normalize_di_lei(r.get("DI_LEI", "")), axis=1)
        df["_di_lei_display"] = df.apply(lambda r: _display_di_lei(r.get("_raw_DI_LEI", r.get("DI_LEI", ""))) or _display_di_lei(r.get("DI_LEI", "")), axis=1)
        df["_di_lei_b5"] = df.apply(lambda r: _b5_di_lei(r.get("_raw_DI_LEI", r.get("DI_LEI", ""))) or _b5_di_lei(r.get("DI_LEI", "")), axis=1)
    else:
        df["_di_lei_norm"] = "其他林地"
        df["_di_lei_display"] = "其他林地"
        df["_di_lei_b5"] = "其他林地"

    lin_zhong_blank = df.apply(_is_blank_lin_zhong, axis=1) if "LIN_ZHONG" in df.columns else pd.Series(False, index=df.index)
    if "_raw_DI_LEI" in df.columns:
        di_lei_codes = df["_raw_DI_LEI"].apply(lambda v: _norm_code(v, 6))
    elif "DI_LEI" in df.columns:
        di_lei_codes = df["DI_LEI"].apply(lambda v: _norm_code(v, 6))
    else:
        di_lei_codes = pd.Series("", index=df.index)
    blank_miao_pu = lin_zhong_blank & (df["_di_lei_norm"] == "苗圃地")
    blank_cai_fa = lin_zhong_blank & (df["_di_lei_norm"] == "采伐迹地")
    blank_030100 = lin_zhong_blank & (di_lei_codes == "030100")
    df.loc[blank_miao_pu, "_land_use"] = "苗圃地"
    df.loc[blank_miao_pu, "_di_lei_norm"] = "苗圃地"
    df.loc[blank_cai_fa, "_land_use"] = "用材林林地"
    df.loc[blank_030100, "_land_use"] = "其他林地"

    if "BH_DJ" in df.columns:
        df["_bh_dj_norm"] = df["BH_DJ"].apply(_normalize_protect_grade)
        df["_bh_dj_display"] = df["BH_DJ"].apply(_display_protect_grade)
    else:
        df["_bh_dj_norm"] = ""
        df["_bh_dj_display"] = ""

    df["_origin_display"] = df.apply(
        lambda r: _display_origin(r.get("_raw_QI_YUAN", r.get("QI_YUAN", "")), r.get("QI_YUAN", "")),
        axis=1
    )
    df["_ling_zu_display"] = df.apply(
        lambda r: _display_lingzu(r.get("LING_ZU", ""), r.get("_di_lei_norm", "") == "竹林地"),
        axis=1
    )

    def _forest_row(r):
        sen = r.get("_raw_SEN_LIN_LB", r.get("SEN_LIN_LB", ""))
        sqdj = r.get("_raw_Y_SQDJ", r.get("Y_SQDJ", ""))
        gj = r.get("_raw_GJGYL_BHDJ", r.get("GJGYL_BHDJ", ""))
        leaf = _forest_leaf(sen, sqdj, gj)
        return pd.Series({
            "_forest_display": _forest_display(sen, sqdj, gj),
            "_forest_leaf": leaf,
            "_forest_b5": _b5_forest_row(leaf),
        })
    df = pd.concat([df, df.apply(_forest_row, axis=1)], axis=1)

    # ---- 重点生态区域（仅ZRBHQ_MC） ----
    def _pa(r):
        mc = r.get("ZRBHQ_MC","")
        if not _has_value(mc): return None
        return _text(mc)
    df["_protected"] = df.apply(_pa, axis=1)

    # ---- 株数 ----
    def _stem(r):
        return _num(r.get("JJLZS", 0))
    df["_stem_count"] = df.apply(_stem, axis=1)

    # 数值化
    for col in ["XBMJ","XIAO_BAN_X","HUO_LMGQXJ","PINGJUN_SG","PINGJUN_XJ","YU_BI_DU","MEI_GQ_ZS","JJLZS","CS_GHQ"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ---- A1_COL_MAP ----
    # 顺序: ZZY字段, A1列号, 转换方式
    # 转换方式: None=直接取值; "xian_name"等特殊名
    col_entries = [
        ("XIAN", 1, "xian_name"),
        ("XIANG", 2, "xiang_name"),
        ("CUN", 3, "cun_name"),
        ("SYLDDKXH", 4, None),
        ("XIAO_BAN", 5, None),
        ("XBMJ", 6, None),
        ("DI_LEI", 7, "_di_lei_display"),
        ("LD_QS", 8, "_ownership"),  # 用我们自己判断的权属
        ("BH_DJ", 9, "_bh_dj_display"),
        ("SEN_LIN_LB", 10, "_forest_display"),
        ("LIN_ZHONG", 11, "_land_use"),  # 使用林地类型(派生)
        ("LIN_ZHONG", 12, "LIN_ZHONG"),  # 林种(原始翻译)
        ("QI_YUAN", 13, "_origin_display"),
        ("protected", 14, "_protected"),  # 重点生态区域
        ("YOU_SHI_SZ", 15, "YOU_SHI_SZ"),
        ("LING_ZU", 16, "_ling_zu_display"),
        ("PINGJUN_SG", 17, None),
        ("PINGJUN_XJ", 18, None),
        ("YU_BI_DU", 19, None),
        ("XIAO_BAN_X", 20, None),  # 小班蓄积
        ("_stem_count", 21, None),
        ("JSNR", 22, None),
        ("LD_XZ", 23, None),
        ("BZ", 24, None),
    ]

    rows = []
    for _, r in df.iterrows():
        row = {}
        for f, ci, tr in col_entries:
            if tr == "xian_name": val = r.get("_xian_name","")
            elif tr == "xiang_name": val = r.get("_xiang_name","")
            elif tr == "cun_name": val = r.get("_cun_name","")
            elif tr == "_land_use": val = r.get("_land_use","")
            elif tr == "_di_lei_display": val = r.get("_di_lei_display","")
            elif tr == "_bh_dj_display": val = r.get("_bh_dj_display","")
            elif tr == "_forest_display": val = r.get("_forest_display","")
            elif tr == "_origin_display": val = r.get("_origin_display","")
            elif tr == "_ling_zu_display": val = r.get("_ling_zu_display","")
            elif tr == "_protected": val = r.get("_protected","") or ""
            elif tr == "_ownership": val = r.get("_ownership","集体")
            elif tr is None: val = r.get(f,"")
            elif tr in dicts:  # 查字典翻译
                d = dicts[tr]
                raw = r.get(f,"")
                val = d.get(str(raw).strip(), raw) if str(raw).strip() in d else raw
            else:
                val = r.get(f,"")
            row[ci] = val
        rows.append(row)

    out = pd.DataFrame(rows)
    for c in [6,20,21]:  # 面积/蓄积/株数列
        if c in out.columns: out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)

    stat = pd.DataFrame({
        "县": df.get("_xian_name", ""),
        "乡": df.get("_xiang_name", ""),
        "村": df.get("_cun_name", ""),
        "面积": pd.to_numeric(df.get("XBMJ", 0), errors="coerce").fillna(0),
        "蓄积": pd.to_numeric(df.get("XIAO_BAN_X", 0), errors="coerce").fillna(0),
        "林地权属": df.get("_ownership", "集体"),
        "使用林地类型": df.get("_land_use", "其他林地"),
        "地类": df.get("_di_lei_norm", "其他林地"),
        "B5地类": df.get("_di_lei_b5", "其他林地"),
        "林地保护等级": df.get("_bh_dj_norm", ""),
        "森林类别": df.get("_forest_display", ""),
        "森林类别叶级": df.get("_forest_leaf", ""),
        "B5森林类别": df.get("_forest_b5", ""),
        "城市规划区": df.get("CS_GHQ", 0),
        "保护区名称": df.get("ZRBHQ_MC", ""),
        "保护区等级": df.get("ZRBHQ_DJ", ""),
        "保护区类型": df.get("ZRBHDLX", ""),
    })
    return out, stat

# ============================================================
# 写 A.1
# ============================================================
def write_a1(ws, a1_data):
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 3:
            try: ws.unmerge_cells(str(mr))
            except: pass
    n = len(a1_data)
    _prepare_plain_rows(ws, 3, n, pattern_row=3, start_col=1, end_col=24)
    for i in range(n):
        r_idx = 3 + i
        for ci in range(1, 25):
            v = a1_data.iloc[i].get(ci)
            if v is None or (isinstance(v, float) and np.isnan(v)): continue
            if isinstance(v, np.integer): v = int(v)
            elif isinstance(v, np.floating): v = round(float(v), 4)
            elif isinstance(v, str) and v.lower() == "nan": continue
            try: ws.cell(row=r_idx, column=ci).value = v
            except: pass
    return n

# ============================================================
# 从 A.1 回读
# ============================================================
def read_a1(ws, nrows):
    hdrs = {}
    for c in range(1, 25): hdrs[c] = ws.cell(2, c).value
    data = []
    for r in range(3, 3 + nrows):
        rd = {}
        has = False
        for c in range(1, 25):
            v = ws.cell(r, c).value
            if v is not None: has = True
            rd[hdrs[c]] = v
        if has and any(rd.values()): data.append(rd)
    if not data: return pd.DataFrame()
    df = pd.DataFrame(data)
    ac = find_col(df, "面积")
    vc = find_col(df, "蓄积")
    if ac: df[ac] = pd.to_numeric(df[ac], errors="coerce").fillna(0)
    if vc: df[vc] = pd.to_numeric(df[vc], errors="coerce").fillna(0)
    return df

# ============================================================
# 构建层级
# ============================================================
def build_hierarchy(a1_df):
    ac = find_col(a1_df, "面积")
    vc = find_col(a1_df, "蓄积")
    cc = find_col(a1_df, "县")
    tc = find_col(a1_df, "乡")
    vlc = find_col(a1_df, "村")
    qc = find_col(a1_df, "权属")
    if not cc: return [], None, None, None
    name = a1_df.iloc[0].get(cc, "县")
    hier = [("合计","total",a1_df,cc,tc,vlc,qc),
            (_subtotal_label(name),"county",a1_df,cc,tc,vlc,qc)]
    if tc and tc in a1_df.columns:
        for tn, tg in a1_df.groupby(tc, sort=False):
            if not _has_value(tn):
                continue
            hier.append((_subtotal_label(tn),"township",tg,cc,tc,vlc,qc))
            if vlc and vlc in tg.columns:
                for vn, vg in tg.groupby(vlc, sort=False):
                    if not _has_value(vn):
                        continue
                    hier.append((_text(vn),"village",vg,cc,tc,vlc,qc))
    return hier, ac, vc, qc

# ============================================================
# B.1 更新
# ============================================================
def update_b1(ws, a1_df):
    ac = find_col(a1_df, "面积")
    vc = find_col(a1_df, "蓄积")
    luc = find_col(a1_df, "使用林地类型")
    if ac is None or luc is None: return
    cat_map = {
        "防护林林地": (5,6), "特用林林地": (7,8), "用材林林地": (9,10),
        "经济林林地": (11,12), "能源林林地": (13,14), "苗圃地": (15,16), "其他林地": (17,18),
    }
    hier, a, v, q = build_hierarchy(a1_df)
    if not hier: return
    _prepare_dynamic_rows(ws, 7, len(hier) * 3, pattern_rows=3)
    _write_b1(ws, hier, ac, vc, q, 7, luc, cat_map)

def _write_b1(ws, hier, ac, vc, qc, start, cat_col, cat_map):
    cur = start
    for en, et, df, cc, tc, vlc, q in hier:
        qu = qc or q
        cd = {}
        if cat_col in df.columns:
            for cn, cp in cat_map.items():
                m = df[cat_col] == cn
                s = df[m]
                if len(s) > 0:
                    ta = s[ac].sum(); tv = s[vc].sum() if vc else 0
                    ga = s.loc[s[qu]=="国有", ac].sum() if qu and qu in s.columns else 0
                    gv = s.loc[s[qu]=="国有", vc].sum() if vc and qu and qu in s.columns else 0
                    cd[cn] = {"ta":ta,"tv":tv,"ga":ga,"gv":gv,"ja":ta-ga,"jv":tv-gv}
        ta = sum(v["ta"] for v in cd.values()); tv = sum(v["tv"] for v in cd.values())
        gt = sum(v["ga"] for v in cd.values()); gtv = sum(v["gv"] for v in cd.values())
        jt = sum(v["ja"] for v in cd.values()); jtv = sum(v["jv"] for v in cd.values())

        label = "合计" if et == "total" else "小计"
        _write_value(ws, cur, 1, en)
        _write_value(ws, cur, 2, label, allow_zero=True)
        _write_value(ws, cur, 3, ta)
        if vc: _write_value(ws, cur, 4, tv)
        for cn, cp in cat_map.items():
            if cn in cd:
                _write_value(ws, cur, cp[0], cd[cn]["ta"])
                if vc: _write_value(ws, cur, cp[1], cd[cn]["tv"])

        _write_value(ws, cur+1, 2, "国有", allow_zero=True)
        _write_value(ws, cur+1, 3, gt)
        if vc: _write_value(ws, cur+1, 4, gtv)
        for cn, cp in cat_map.items():
            if cn in cd:
                _write_value(ws, cur+1, cp[0], cd[cn]["ga"])
                if vc: _write_value(ws, cur+1, cp[1], cd[cn]["gv"])

        _write_value(ws, cur+2, 2, "集体", allow_zero=True)
        _write_value(ws, cur+2, 3, jt)
        if vc: _write_value(ws, cur+2, 4, jtv)
        for cn, cp in cat_map.items():
            if cn in cd:
                _write_value(ws, cur+2, cp[0], cd[cn]["ja"])
                if vc: _write_value(ws, cur+2, cp[1], cd[cn]["jv"])

        try: ws.merge_cells(start_row=cur, start_column=1, end_row=cur+2, end_column=1)
        except: pass
        cur += 3
    return cur

# ============================================================
# B.2/B.4 通用面积sheet
# ============================================================
def _write_area_sheet(ws, a1_df, cat_col, cat_map, start):
    ac = find_col(a1_df, "面积")
    if ac is None or cat_col is None or cat_col not in a1_df.columns: return
    hier, _, _, _ = build_hierarchy(a1_df)
    if not hier: return
    _prepare_dynamic_rows(ws, start, len(hier) * 3, pattern_rows=3)
    qc = find_col(a1_df, "权属")
    cur = start
    for en, et, df, cc, tc, vlc, q in hier:
        qu = qc or q
        total = df[ac].sum()
        _write_value(ws, cur, 1, en)
        _write_value(ws, cur, 2, "合计" if et=="total" else "小计", allow_zero=True)
        _write_value(ws, cur, 3, total)
        for cn, col in cat_map.items():
            a = df.loc[df[cat_col]==cn, ac].sum()
            _write_value(ws, cur, col, a)
        _write_value(ws, cur+1, 2, "国有", allow_zero=True)
        if qu and qu in df.columns:
            _write_value(ws, cur+1, 3, df.loc[df[qu]=="国有", ac].sum())
        for cn, col in cat_map.items():
            a = df.loc[(df[cat_col]==cn)&(df[qu]=="国有"), ac].sum() if qu and qu in df.columns else 0
            _write_value(ws, cur+1, col, a)
        _write_value(ws, cur+2, 2, "集体", allow_zero=True)
        if qu and qu in df.columns:
            _write_value(ws, cur+2, 3, df.loc[df[qu]=="集体", ac].sum())
        for cn, col in cat_map.items():
            a = df.loc[(df[cat_col]==cn)&(df[qu]=="集体"), ac].sum() if qu and qu in df.columns else 0
            _write_value(ws, cur+2, col, a)
        try: ws.merge_cells(start_row=cur, start_column=1, end_row=cur+2, end_column=1)
        except: pass
        cur += 3

def update_b2(ws, a1_df):
    cm = {"乔木林地":4,"竹林地":5,"特殊灌林地":6,"一般灌木林地":7,"疏林地":8,
          "未成林地":9,"苗圃地":10,"采伐迹地":11,"火烧迹地":12,"宜林地":13,"其他林地":14}
    dc = "地类" if "地类" in a1_df.columns else find_col(a1_df, "地类")
    _write_area_sheet(ws, a1_df, dc, cm, 6)

def update_b4(ws, a1_df):
    cm = {"Ⅰ级":4,"Ⅱ级":5,"Ⅲ级":6,"Ⅳ级":7}
    bc = find_col(a1_df, "保护等级")
    _write_area_sheet(ws, a1_df, bc, cm, 6)

# ============================================================
# B.3 更新（按森林类别）
# ============================================================
def update_b3(ws, a1_df):
    ac = find_col(a1_df, "面积")
    leaf_col = find_col(a1_df, "森林类别叶级")
    if ac is None or leaf_col is None: return
    hier, _, _, _ = build_hierarchy(a1_df)
    if not hier: return
    _prepare_dynamic_rows(ws, 8, len(hier) * 3, pattern_rows=3)
    qc = find_col(a1_df, "权属")
    leaf_cols = {
        "国家级一级": 6, "国家级二级": 7, "国家级": 5,
        "省级": 8, "市县级": 9,
        "重点商品": 11, "一般商品": 12,
    }

    def sums(row_df, owner=None):
        if owner and qc and qc in row_df.columns:
            row_df = row_df[row_df[qc] == owner]
        vals = {c: 0 for c in range(3, 13)}
        vals[3] = row_df[ac].sum()
        for leaf, col in leaf_cols.items():
            vals[col] += row_df.loc[row_df[leaf_col] == leaf, ac].sum()
        vals[5] = vals[5] + vals[6] + vals[7]
        vals[4] = vals[5] + vals[8] + vals[9]
        vals[10] = vals[11] + vals[12]
        if vals[3] == 0:
            vals[3] = vals[4] + vals[10]
        return vals

    cur = 8
    for en, et, df, cc, tc, vlc, q in hier:
        label = "合计" if et == "total" else "小计"
        _write_value(ws, cur, 1, en)
        _write_value(ws, cur, 2, label, allow_zero=True)
        vals = sums(df)
        for col, val in vals.items():
            _write_value(ws, cur, col, val)

        _write_value(ws, cur+1, 2, "国有", allow_zero=True)
        vals = sums(df, "国有")
        for col, val in vals.items():
            _write_value(ws, cur+1, col, val)

        _write_value(ws, cur+2, 2, "集体", allow_zero=True)
        vals = sums(df, "集体")
        for col, val in vals.items():
            _write_value(ws, cur+2, col, val)

        try: ws.merge_cells(start_row=cur, start_column=1, end_row=cur+2, end_column=1)
        except: pass
        cur += 3

# ============================================================
# B.5 更新
# ============================================================
def update_b5(ws, a1_df):
    ac = find_col(a1_df, "面积")
    dc = find_col(a1_df, "B5地类")
    fc = find_col(a1_df, "B5森林类别")
    qc = find_col(a1_df, "权属")
    cc = find_col(a1_df, "县")
    if ac is None or dc is None or fc is None or qc is None: return
    _reset_merges(ws, ["A6:A12", "B6:B9", "B10:B12", "A13:A19", "B13:B16", "B17:B19"])
    _clear_values(ws, 6, 19, 1, 13)

    county_name = a1_df.iloc[0].get(cc, "X县") if cc and len(a1_df) else "X县"
    labels = {
        6: ("合计", "公益林地", "小计"),
        7: ("", "", "国家级公益林地"),
        8: ("", "", "省级公益林地"),
        9: ("", "", "其他公益林地"),
        10: ("", "商品林地", "小计"),
        11: ("", "", "重点商品林地"),
        12: ("", "", "一般商品林地"),
        13: (county_name, "公益林地", "小计"),
        14: ("", "", "国家级公益林地"),
        15: ("", "", "省级公益林地"),
        16: ("", "", "其他公益林地"),
        17: ("", "商品林地", "小计"),
        18: ("", "", "重点商品林地"),
        19: ("", "", "一般商品林地"),
    }
    for row, (a, b, c) in labels.items():
        if a: _write_value(ws, row, 1, a, allow_zero=True)
        if b: _write_value(ws, row, 2, b, allow_zero=True)
        if c: _write_value(ws, row, 3, c, allow_zero=True)

    row_defs = {
        7: ["国家级公益林地"],
        8: ["省级公益林地"],
        9: ["其他公益林地"],
        11: ["重点商品林地"],
        12: ["一般商品林地"],
    }
    row_defs[6] = row_defs[7] + row_defs[8] + row_defs[9]
    row_defs[10] = row_defs[11] + row_defs[12]
    for src, dst in [(6, 13), (7, 14), (8, 15), (9, 16), (10, 17), (11, 18), (12, 19)]:
        row_defs[dst] = row_defs[src]

    type_cols = {
        "乔木林地": (6, 7),
        "竹林地": (8, 9),
        "特殊灌木林地": (10, 11),
        "其他林地": (12, 13),
    }

    def write_row(row, cats):
        sub = a1_df[a1_df[fc].isin(cats)]
        gy_total = sub.loc[sub[qc] == "国有", ac].sum()
        jt_total = sub.loc[sub[qc] == "集体", ac].sum()
        _write_value(ws, row, 4, gy_total)
        _write_value(ws, row, 5, jt_total)
        for dl, (gy_col, jt_col) in type_cols.items():
            dsub = sub[sub[dc] == dl]
            _write_value(ws, row, gy_col, dsub.loc[dsub[qc] == "国有", ac].sum())
            _write_value(ws, row, jt_col, dsub.loc[dsub[qc] == "集体", ac].sum())

    for row, cats in row_defs.items():
        write_row(row, cats)

# ============================================================
# B.6 更新（仅ZRBHQ_MC）
# ============================================================
def update_b6(ws, raw_zzy):
    rows = []
    for _, r in raw_zzy.iterrows():
        mc = r.get("保护区名称", r.get("ZRBHQ_MC", ""))
        if not _has_value(mc): continue
        area = _num(r.get("面积", r.get("XBMJ", 0)))
        dj_s = _text(r.get("保护区等级", r.get("ZRBHQ_DJ", "")))
        lx = _text(r.get("保护区类型", r.get("ZRBHDLX", "")))
        rows.append({"name": _text(mc), "level": dj_s, "type": lx, "area": area})
    _prepare_plain_rows(ws, 6, max(len(rows), 1), pattern_row=6, start_col=1, end_col=19)
    if not rows: return  # 无保护区数据，保持空白
    pdf = pd.DataFrame(rows)
    agg = pdf.groupby(["name","level","type"], dropna=False)["area"].sum().reset_index()
    cur = 6
    for _, r in agg.iterrows():
        _write_value(ws, cur, 1, r["name"])
        _write_value(ws, cur, 2, r["area"])
        _write_value(ws, cur, 3, r["level"])
        _write_value(ws, cur, 4, r["area"])
        _write_value(ws, cur, 5, r["type"])
        cur += 1

# ============================================================
# Fee 更新
# ============================================================
def update_fee(ws, a1_df):
    ac = find_col(a1_df, "面积")
    dc = "地类" if "地类" in a1_df.columns else find_col(a1_df, "地类")
    leaf_col = find_col(a1_df, "森林类别叶级")
    cc = find_col(a1_df, "县")
    city_col = find_col(a1_df, "城市规划区")
    if ac is None or dc is None: return
    _write_value(ws, 6, 1, a1_df.iloc[0].get(cc, "县") if cc and len(a1_df)>0 else "县")
    fee_cat = {"乔木林地":"乔木林地（含迹地）","竹林地":"竹林地","苗圃地":"苗圃地",
               "特殊灌林地":"灌木林地（含特灌林地）","一般灌木林地":"灌木林地（含特灌林地）",
               "疏林地":"疏林地","未成林地":"未成林造林地","宜林地":"宜林地",
               "采伐迹地":"乔木林地（含迹地）","火烧迹地":"乔木林地（含迹地）","其他林地":"其他林地"}
    rate_map = {"乔木林地（含迹地）":10,"竹林地":10,"苗圃地":10,
                "灌木林地（含特灌林地）":6,"疏林地":6,"未成林造林地":6,"宜林地":3,"其他林地":6}
    row_map = {}
    for r in range(7, ws.max_row + 1):
        name = _text(ws.cell(r, 1).value)
        if name:
            row_map[name] = r

    work = a1_df.copy()
    work["_fee_cat"] = work[dc].map(fee_cat).fillna("其他林地")
    work["_fee_public"] = work[leaf_col].apply(_fee_public_class) if leaf_col in work.columns else False
    work["_fee_city"] = work[city_col].apply(_is_city_planning) if city_col in work.columns else False

    _clear_values(ws, 6, ws.max_row, 2, 11)
    row_totals = {}
    for _, r in work.iterrows():
        cat = r["_fee_cat"]
        ri = row_map.get(cat)
        if ri is None:
            continue
        area = _num(r.get(ac))
        if area == 0:
            continue
        rate = rate_map.get(cat, 6)
        row_totals.setdefault(ri, {"C": 0, "D": 0, "E": 0, "F": 0, "rate": rate})
        if r["_fee_city"] and r["_fee_public"]:
            row_totals[ri]["C"] += area
        elif r["_fee_city"]:
            row_totals[ri]["D"] += area
        elif r["_fee_public"]:
            row_totals[ri]["E"] += area
        else:
            row_totals[ri]["F"] += area

    grand = {"B": 0, "C": 0, "D": 0, "E": 0, "F": 0, "K": 0}
    for ri, fr in row_totals.items():
        c, d, e, f = fr["C"], fr["D"], fr["E"], fr["F"]
        rate = fr["rate"]
        total_area = c + d + e + f
        fee = c * rate * 4 + d * rate * 2 + e * rate * 2 + f * rate
        _write_value(ws, ri, 2, total_area)
        _write_value(ws, ri, 3, c)
        _write_value(ws, ri, 4, d)
        _write_value(ws, ri, 5, e)
        _write_value(ws, ri, 6, f)
        if c > 0: _write_value(ws, ri, 7, rate * 4)
        if d > 0: _write_value(ws, ri, 8, rate * 2)
        if e > 0: _write_value(ws, ri, 9, rate * 2)
        if f > 0: _write_value(ws, ri, 10, rate)
        _write_value(ws, ri, 11, fee, allow_zero=True)
        grand["B"] += total_area; grand["C"] += c; grand["D"] += d
        grand["E"] += e; grand["F"] += f; grand["K"] += fee

    for col, key in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (11, "K")]:
        _write_value(ws, 6, col, grand[key])

# ============================================================
# 处理单县
# ============================================================
def discover_counties(base_dir=BASE_DIR):
    if not os.path.isdir(base_dir):
        return []
    counties = []
    for cname in sorted(os.listdir(base_dir)):
        county_dir = os.path.join(base_dir, cname)
        zzy = os.path.join(county_dir, "林地图斑", "ZZY.shp")
        tpl = os.path.join(county_dir, "可研数据", "可研附表模板.xlsx")
        if os.path.isdir(county_dir) and os.path.exists(zzy) and os.path.exists(tpl):
            counties.append(cname)
    return counties


def process_county(cname, dicts, c2n, town_names, vill_names):
    print(f"\n===== {cname} =====")
    d = os.path.join(BASE_DIR, cname)
    zzy = os.path.join(d, "林地图斑", "ZZY.shp")
    tpl = os.path.join(d, "可研数据", "可研附表模板.xlsx")
    if not os.path.exists(zzy): print("  跳过: 无ZZY.shp"); return False
    if not os.path.exists(tpl): print("  跳过: 无模板"); return False

    gdf = read_zzy_shp(zzy)
    if gdf is None or len(gdf)==0: print("  跳过: 无数据"); return False
    print(f"  ZZY: {len(gdf)} 条")

    a1d, stat_df = prepare_a1_data(gdf, dicts, c2n, town_names, vill_names)
    print(f"  A.1数据: {len(a1d)} 行")

    wb = load_workbook(tpl)
    n = write_a1(wb[SHEET_A1], a1d)
    print(f"  写入A.1: {n} 行")

    if stat_df.empty: print("  跳过B-sheet"); wb.save(tpl); return True

    update_b1(wb[SHEET_B1], stat_df)
    update_b2(wb[SHEET_B2], stat_df)
    update_b3(wb[SHEET_B3], stat_df)
    update_b4(wb[SHEET_B4], stat_df)
    update_b5(wb[SHEET_B5], stat_df)
    update_b6(wb[SHEET_B6], stat_df)
    update_fee(wb[SHEET_FEE], stat_df)
    wb.save(tpl)
    print("  ✓")
    return True

def zip_county_dirs(counties):
    print("\n[4] 压缩县目录...")
    zip_paths = []
    for cname in counties:
        county_dir = os.path.join(BASE_DIR, cname)
        if not os.path.isdir(county_dir):
            print(f"  跳过: 无目录 {cname}")
            continue

        zip_path = os.path.join(BASE_DIR, f"{cname}.zip")
        tmp_path = zip_path + ".tmp"
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

        added = 0
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, _, files in os.walk(county_dir):
                files.sort()
                for fn in files:
                    fp = os.path.join(root, fn)
                    zf.write(fp, os.path.relpath(fp, BASE_DIR))
                    added += 1

        if os.path.exists(zip_path):
            os.remove(zip_path)
        os.replace(tmp_path, zip_path)
        zip_paths.append(zip_path)
        print(f"  zip: {zip_path} ({added} 个文件)")
    return zip_paths

# ============================================================
# Main
# ============================================================
def main():
    print("="*50); print("可研附表模板自动填充"); print("="*50)
    print("\n[1] 解析字典...")
    dicts = parse_md_dictionaries(MD_FILE)
    for n, d in dicts.items(): print(f"  {n}: {len(d)} 项")

    print("\n  解析乡镇/村名称...")
    tn = parse_township_names(MD_FILE)
    vn = parse_village_names(MD_FILE)
    print(f"  乡镇: {len(tn)} 项, 行政村: {len(vn)} 项")

    print("\n[2] 县名...")
    c2n = dict(COUNTY_CODE_TO_NAME)
    print(f"  {len(c2n)} 个县")

    print("\n[3] 处理...")
    counties = discover_counties()
    if not counties:
        print(f"  未发现可处理县目录: {BASE_DIR}")
        return
    print(f"  发现 {len(counties)} 个县: {', '.join(counties)}")
    ok = sum(1 for c in counties if process_county(c, dicts, c2n, tn, vn))
    print(f"\n完成: {ok}/{len(counties)}")
    zip_county_dirs(counties)

if __name__ == "__main__":
    main()
