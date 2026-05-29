import openpyxl
import sys

path = r"c:\4code\3lot\模版-1009征占用林地数据模板CGCG2000_111\征占shp文件数据结构20231009.xlsx"

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheet names:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== {name} ===")
        for row in ws.iter_rows(values_only=False):
            vals = [str(c.value) if c.value is not None else "" for c in row]
            print("|".join(vals))
except Exception as e:
    print(f"Error: {e}", file=sys.stderr)
    sys.exit(1)
