# -*- coding: utf-8 -*-
"""
按县生成报告小任务结果。

每个小任务写成一个独立函数，并在 TASKS 中注册。
运行时通过 --task 选择任务，结果输出为 Markdown。
"""

import argparse
import importlib.util
from pathlib import Path

from project_config import OUTPUT_BASE, STANDARD_FILE, COUNTY_CODE_TO_NAME


PROJECT_NAME = "湖南省洞庭湖区重点垸堤防加固二期工程"
PROJECT_USE = "主要用于一线防洪大堤堤防加固等建设用地"
DEFAULT_OUTPUT = Path(__file__).with_name("6-报告小任务结果.md")

_PROTECTED_AREA_FIELDS = [
    "ZRBHQ_MC", "SLGY_MC", "SDGY_MC", "FJMSQ_MC",
    "DZGY_MC", "HYGY_MC", "SMGY_MC", "CYGY_MC",
]

# 可研附表 B.1 使用林地类型固定顺序
_B1_LAND_USE_ORDER = [
    "防护林林地", "特用林林地", "用材林林地", "经济林林地",
    "能源林林地", "苗圃地", "其他林地",
]

_dicts_cache = None


def load_code5():
    path = Path(__file__).with_name("5-populate_template.py")
    spec = importlib.util.spec_from_file_location("populate_template_code5", str(path))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def county_name(code5, value, fallback):
    code = code5._norm_code(value, 6)
    return COUNTY_CODE_TO_NAME.get(code) or fallback


def location_text(df, code5, town_names, vill_names, fallback_county):
    groups = location_groups(df, code5, town_names, vill_names, fallback_county)

    by_town = []
    for town, village in groups:
        if not by_town or by_town[-1][0] != town:
            by_town.append([town, []])
        by_town[-1][1].append(village)

    parts = []
    for town, villages in by_town:
        clean_villages = [v for v in villages if v and v != town]
        parts.append("{}{}".format(town, "、".join(clean_villages)))
    return "、".join(parts)


def location_groups(df, code5, town_names, vill_names, fallback_county):
    seen = set()
    groups = []

    for _, row in df.iterrows():
        xian_code = code5._norm_code(row.get("XIAN", ""), 6)
        xiang_code = code5._norm_code(row.get("XIANG", ""), 3)
        cun_code = code5._norm_code(row.get("CUN", ""), 3)
        xian_name = COUNTY_CODE_TO_NAME.get(xian_code) or fallback_county
        town = town_names.get((xian_code, xiang_code)) or xiang_code
        village = (
            vill_names.get((xian_name, xiang_code, cun_code))
            or vill_names.get((xiang_code, cun_code))
            or cun_code
        )
        key = (town, village)
        if town and village and key not in seen:
            seen.add(key)
            groups.append(key)
    return groups


def format_area(area):
    return f"{area:.4f}".rstrip("0").rstrip(".")


def task_forestland_application(cname, df, code5, town_names, vill_names):
    """使用林地申请说明：按县生成一段文字。"""
    area = sum(code5._num(v) for v in df.get("XBMJ", []))
    first_xian = df.iloc[0].get("XIAN", "") if len(df) else ""
    county = county_name(code5, first_xian, cname)
    location = location_text(df, code5, town_names, vill_names, county)

    return (
        f"{PROJECT_NAME}需使用{county}林地{format_area(area)}公顷，"
        f"{PROJECT_USE}，地点位于{location}。林地权属界线明晰，无争议。\n"
        f"为确保项目的顺利实施，特申请项目使用林地{format_area(area)}公顷。"
    )


def ownership_detail_line(label, df, code5, town_names, vill_names, county):
    if df.empty:
        return f"{label}林地：无。"

    location = location_text(df, code5, town_names, vill_names, county)
    groups = location_groups(df, code5, town_names, vill_names, county)
    town_count = len({town for town, _village in groups})
    village_count = len({(town, village) for town, village in groups})
    return (
        f"{label}林地：{county}等1个县（区、市）"
        f"{location}等{town_count}个镇{village_count}个行政村。"
    )


def task_used_forestland_unit_detail(cname, df, code5, town_names, vill_names):
    """被使用林地单位明细表：按国有、集体林地列出涉及单位。"""
    first_xian = df.iloc[0].get("XIAN", "") if len(df) else ""
    county = county_name(code5, first_xian, cname)

    if "LD_QS" in df.columns:
        ownership = df["LD_QS"].apply(code5._normalize_ownership)
    else:
        ownership = "集体"

    state_df = df.loc[ownership == "国有"]
    collective_df = df.loc[ownership == "集体"]
    return "\n".join([
        "被使用林地单位明细表",
        ownership_detail_line("国有", state_df, code5, town_names, vill_names, county),
        ownership_detail_line("集体", collective_df, code5, town_names, vill_names, county),
    ])


def task_natural_forest_area(cname, df, code5, town_names, vill_names):
    """天然林林地面积：按县统计起源(QI_YUAN)为天然的拟使用林地面积。"""
    first_xian = df.iloc[0].get("XIAN", "") if len(df) else ""
    county = county_name(code5, first_xian, cname)

    total = sum(code5._num(v) for v in df.get("XBMJ", []))
    natural = 0.0
    if "QI_YUAN" in df.columns:
        for _, row in df.iterrows():
            if code5._display_origin(row.get("QI_YUAN", "")) == "天然":
                natural += code5._num(row.get("XBMJ", 0))

    if total > 0:
        return (
            f"{county}天然林林地面积{format_area(natural)}公顷，"
            f"占拟使用林地面积{format_area(total)}公顷的{natural / total * 100:.2f}%。"
        )
    return f"{county}天然林林地面积{format_area(natural)}公顷。"


def _load_dicts(code5):
    global _dicts_cache
    if _dicts_cache is None:
        _dicts_cache = code5.parse_md_dictionaries(STANDARD_FILE)
    return _dicts_cache


def _area4(x):
    return "{:.4f}".format(float(x))


def _read_xmhx_total_area(code5, cname):
    """读取 项目红线/XMHX.shp 的 XMPFNYTDMJ（项目用地总面积，公顷）。"""
    xmhx = Path(OUTPUT_BASE) / cname / "项目红线" / "XMHX.shp"
    df = code5.read_zzy_shp(str(xmhx))
    if df is None or len(df) == 0 or "XMPFNYTDMJ" not in df.columns:
        return None
    return sum(code5._num(v) for v in df["XMPFNYTDMJ"])


def _b1_land_use(code5, dicts, r):
    """复刻可研附表 B.1 的使用林地类型判定（5号脚本617-645行）。"""
    raw = r.get("_raw_LIN_ZHONG", r.get("LIN_ZHONG", ""))
    lu = code5._normalize_land_use(raw)
    if not code5._has_value(raw):
        di_norm = code5._normalize_di_lei(r.get("_raw_DI_LEI", r.get("DI_LEI", "")))
        di_code = code5._norm_code(r.get("_raw_DI_LEI", r.get("DI_LEI", "")), 6)
        if di_norm == "苗圃地":
            lu = "苗圃地"
        elif di_norm == "采伐迹地":
            lu = "用材林林地"
        elif di_code == "030100":
            lu = "其他林地"
    return lu


def _category_line(df, code5, name_fn, label, order=None):
    """按 name_fn(row) 分组汇总 XBMJ，生成“{label}：A0.1234公顷、B0.5678公顷。”"""
    groups = {}
    for _, r in df.iterrows():
        nm = name_fn(r)
        if not nm:
            continue
        groups[nm] = groups.get(nm, 0.0) + code5._num(r.get("XBMJ", 0))
    if not groups:
        return "{}：无。".format(label)
    if order:
        names = [n for n in order if n in groups]
        items = [(n, groups[n]) for n in names]
    else:
        items = sorted(groups.items(), key=lambda kv: -kv[1])
    body = "、".join("{}{}公顷".format(nm, _area4(a)) for nm, a in items)
    return "{}：{}。".format(label, body)


def _vegetation_line(df, code5, dicts):
    species = {}
    for _, r in df.iterrows():
        code = code5._norm_code(r.get("YOU_SHI_SZ", ""), 6)
        if not code:
            continue
        nm = dicts.get("YOU_SHI_SZ", {}).get(code, "")
        if not nm or nm.startswith("其它") or nm.startswith("其他"):
            continue
        species[nm] = species.get(nm, 0.0) + code5._num(r.get("XBMJ", 0))
    if not species:
        return "项目区现状植被以【植被待填】为主。"
    items = sorted(species.items(), key=lambda kv: -kv[1])
    names = "、".join(nm for nm, _ in items)
    return "项目区现状植被以{}为主。".format(names)


def _discrepancy_line(df, code5, dicts, county):
    """BZ=实地调查 的小班：原图地类名称取 DLBM 名，原图地类为非林地，
    现场调查地类取当前 DI_LEI 名（DI_LEI 已被改写为纠正后的林地类型）。"""
    dlbm_dict = dicts.get("DLBM", {})
    dilei_dict = dicts.get("DI_LEI", {})
    groups = {}
    for _, r in df.iterrows():
        if code5._text(r.get("BZ", "")) != "实地调查":
            continue
        dl = code5._norm_code(r.get("DLBM", ""), 4)
        dl_name = dlbm_dict.get(dl, "") or "其他林地"
        di = code5._norm_code(r.get("DI_LEI", ""), 6)
        di_name = dilei_dict.get(di, "") or code5._display_di_lei(r.get("DI_LEI", ""))
        key = (dl_name, di_name)
        g = groups.setdefault(key, {"seqs": [], "area": 0.0})
        try:
            g["seqs"].append(int(float(r.get("SYLDDKXH", 0) or 0)))
        except (ValueError, TypeError):
            continue
        g["area"] += code5._num(r.get("XBMJ", 0))

    total_area = sum(code5._num(v) for v in df.get("XBMJ", []))
    n_plots = len(df)
    head = "项目拟使用林地{}公顷，共{}个小班".format(_area4(total_area), n_plots)
    tail = "项目区其它拟使用林地属性因子与{}2022年林草湿资源一张图的属性因子一致。".format(county)
    if not groups:
        return head + "，" + tail

    parts = []
    for (dl_name, di_name), g in sorted(groups.items(), key=lambda kv: min(kv[1]["seqs"])):
        seqs = sorted(g["seqs"])
        seq_txt = "、".join(str(s) for s in seqs)
        cnt = len(seqs)
        if cnt == 1:
            seg = "使用林地地块序号{}号小班".format(seq_txt)
        else:
            seg = "使用林地地块序号{}号等{}个小班".format(seq_txt, cnt)
        parts.append(
            "{}，面积{}公顷，在{}2022年林草湿资源一张图中地类名称为{}，"
            "地类为非林地，根据追溯和现场调查地类为{}".format(
                seg, _area4(g["area"]), county, dl_name, di_name
            )
        )
    return head + "，其中" + "；".join(parts) + "。" + tail


def _protected_area_line(df, code5, county):
    # 按保护区名称分组，记录涉及的小班序号与面积
    pa_groups = {}
    for _, r in df.iterrows():
        for f in _PROTECTED_AREA_FIELDS:
            if f not in df.columns:
                continue
            nm = code5._text(r.get(f, ""))
            if not nm:
                continue
            g = pa_groups.setdefault(nm, {"seqs": [], "area": 0.0})
            try:
                seq = int(float(r.get("SYLDDKXH", 0) or 0))
                if seq > 0:
                    g["seqs"].append(seq)
            except (ValueError, TypeError):
                pass
            g["area"] += code5._num(r.get("XBMJ", 0))
            break

    if pa_groups:
        parts = []
        for nm, g in sorted(pa_groups.items(), key=lambda kv: min(kv[1]["seqs"])):
            seqs = sorted(g["seqs"])
            seq_txt = "、".join(str(s) for s in seqs)
            cnt = len(seqs)
            if cnt == 1:
                seg = "使用林地地块序号{}号小班".format(seq_txt)
            else:
                seg = "使用林地地块序号{}号等{}个小班".format(seq_txt, cnt)
            parts.append("{}，面积{}公顷，位于{}".format(seg, _area4(g["area"]), nm))
        pa_part = "项目区涉及重点保护区域，其中" + "；".join(parts) + "。未涉及重要水源保护区和{}生态保护红线等生态敏感区域。".format(county)
    else:
        pa_part = (
            "项目区未涉及国家公园、自然保护区、森林公园、湿地公园、风景名胜区、"
            "世界遗产以及其他自然公园等重点保护区域，未涉及重要水源保护区和"
            "{}生态保护红线等生态敏感区域。".format(county)
        )
    wildlife = (
        "项目区及周边没有国家级和省级重点保护的野生植物和古树名木，"
        "也没有国家级和省级重点保护的野生动物及其栖息地。"
    )
    in_city = any(code5._is_city_planning(v) for v in df.get("CS_GHQ", [])) if "CS_GHQ" in df.columns else False
    if in_city:
        city_part = "项目区拟使用林地位于{}城市规划区范围内。".format(county)
    else:
        city_part = "项目区拟使用林地不在{}城市规划区范围内。".format(county)
    return pa_part + wildlife + city_part


def _protect_grades_text(df, code5):
    """列出涉及的全部保护等级，按 Ⅰ→Ⅳ 升序，如“Ⅱ级、Ⅲ级、Ⅳ级保护林地”。"""
    rank = ["Ⅰ级", "Ⅱ级", "Ⅲ级", "Ⅳ级", "Ⅴ级"]
    present = set()
    for _, r in df.iterrows():
        g = code5._normalize_protect_grade(r.get("BH_DJ", ""))
        if g:
            present.add(g)
    ordered = [g for g in rank if g in present]
    others = sorted(present - set(rank))
    return "、".join(ordered + others) + "保护林地"


def _ownership_summary(df, code5):
    """权属描述：仅集体→“全为集体”；含国有→“为集体和国有”。"""
    types = set()
    for _, r in df.iterrows():
        types.add(code5._normalize_ownership(r.get("LD_QS", "")))
    if types == {"集体"}:
        return "全为集体"
    if types == {"国有"}:
        return "全为国有"
    return "为集体和国有"


def _volume_sum(df, code5):
    """需采伐林木蓄积量（立方米）：小班蓄积 XIAO_BAN_X 合计，取整。"""
    return int(round(sum(code5._num(v) for v in df.get("XIAO_BAN_X", []))))


def task_review_opinion(cname, df, code5, town_names, vill_names):
    """审查意见：按县生成完整审查意见报告。"""
    dicts = _load_dicts(code5)
    first_xian = df.iloc[0].get("XIAN", "") if len(df) else ""
    county = county_name(code5, first_xian, cname)

    total_area = sum(code5._num(v) for v in df.get("XBMJ", []))
    location = location_text(df, code5, town_names, vill_names, county)
    xz = code5._text(df.iloc[0].get("LD_XZ", "")) if len(df) else ""
    if not xz:
        xz = "永久占用"
    yongdi = _read_xmhx_total_area(code5, cname)
    yongdi_txt = _area4(yongdi) if yongdi is not None else "【用地面积待填】"
    volume = _volume_sum(df, code5)

    natural = 0.0
    for _, r in df.iterrows():
        if code5._display_origin(r.get("QI_YUAN", "")) == "天然":
            natural += code5._num(r.get("XBMJ", 0))

    lines = []
    lines.append("项目区在{}拟用地总面积{}公顷，其中拟使用林地面积{}公顷。".format(
        county, yongdi_txt, _area4(total_area)))

    lines.append("二、拟使用林地和采伐林木情况")
    lines.append("项目区拟使用林地位于{}{}，拟使用林地位置、范围与现地一致。"
                 "项目区拟使用林地性质为{}，拟使用林地面积{}公顷，权属清晰，无山林权属纠纷。".format(
                     county, location, xz, _area4(total_area)))
    lines.append(_category_line(df, code5, lambda r: _b1_land_use(code5, dicts, r),
                                "按使用林地类型分", order=_B1_LAND_USE_ORDER))
    lines.append(_category_line(df, code5,
                                lambda r: dicts.get("DI_LEI", {}).get(
                                    code5._norm_code(r.get("DI_LEI", ""), 6), "")
                                or code5._display_di_lei(r.get("DI_LEI", "")),
                                "按地类分"))
    lines.append(_category_line(df, code5,
                                lambda r: code5._forest_display(
                                    r.get("SEN_LIN_LB", ""), r.get("Y_SQDJ", ""), r.get("GJGYL_BHDJ", "")),
                                "按森林类别分"))
    lines.append(_category_line(df, code5,
                                lambda r: "{}保护林地".format(code5._normalize_protect_grade(r.get("BH_DJ", "")))
                                if code5._normalize_protect_grade(r.get("BH_DJ", "")) else "",
                                "按林地保护等级分"))
    lines.append(_category_line(df, code5,
                                lambda r: "{}林地".format(code5._normalize_ownership(r.get("LD_QS", ""))),
                                "按林地权属分"))
    lines.append("项目区涉及天然林{}公顷。".format(_area4(natural)) if natural > 0 else "未涉及天然林。")
    lines.append("项目区拟使用林地需采伐林木的蓄积量为{}立方米。".format(volume))

    lines.append("三、项目建设符合林地保护利用规划和使用林地的条件")
    lines.append("项目是经国家发展和改革委员会批复的基础设施项目，可使用Ⅱ级保护林地。"
                 "项目区拟使用林地涉及{}，符合林地等级管理规定要求，符合使用林地的条件和范围。".format(
                     _protect_grades_text(df, code5)))

    lines.append("四、使用林地定额情况")
    lines.append("项目区拟使用林地面积{}公顷，占用湖南省当年定额指标。".format(_area4(total_area)))

    lines.append("五、林草湿资源“一张图”情况")
    lines.append(_discrepancy_line(df, code5, dicts, county))

    lines.append("六、现场查验情况")
    lines.append("2026年  月  日，我局组织技术人员对该项目使用林地进行了现场查验。")
    lines.append("项目拟使用林地位置，范围与现场核实一致；" + _protected_area_line(df, code5, county))
    lines.append("项目拟使用林地四至界限清晰，林地权属{}、无争议。"
                 "项目区拟使用林地不存在未批先占或擅自改变林地用途、采伐林木的行为。".format(
                     _ownership_summary(df, code5)))

    lines.append("七、公示情况")
    lines.append("2026年  月  日至2026年  月  日，该项目由{}林业局在{}{}进行了公示，"
                 "在公示期内无异议，也没有收到第三方任何意见反馈。".format(county, county, location))

    lines.append("综上所述，该项目符合《建设项目使用林地审核审批管理办法》"
                 "（国家林业局令第35号）的有关规定，我局拟同意该项目使用林地申请。")
    return "\n".join(lines)


def task_forestland_overview(cname, df, code5, town_names, vill_names):
    """使用林地情况说明：按县生成完整说明文字。"""
    dicts = _load_dicts(code5)
    first_xian = df.iloc[0].get("XIAN", "") if len(df) else ""
    county = county_name(code5, first_xian, cname)

    total_area = sum(code5._num(v) for v in df.get("XBMJ", []))
    location = location_text(df, code5, town_names, vill_names, county)
    xz = code5._text(df.iloc[0].get("LD_XZ", "")) if len(df) else ""
    if not xz:
        xz = "永久占用"
    yongdi = _read_xmhx_total_area(code5, cname)
    yongdi_txt = _area4(yongdi) if yongdi is not None else "【用地面积待填】"

    paragraphs = [
        "项目区拟使用林地位于{}{}，项目区拟使用林地位置、范围与现地一致。"
        "项目区拟使用林地性质为{}。".format(county, location, xz),
        "项目区在{}用地面积{}公顷，其中拟使用林地面积{}公顷。"
        "项目区拟使用林地{}。".format(county, yongdi_txt, _area4(total_area), PROJECT_USE),
        _category_line(
            df, code5,
            lambda r: _b1_land_use(code5, dicts, r),
            "按使用林地类型分",
            order=_B1_LAND_USE_ORDER,
        ),
        _category_line(
            df, code5,
            lambda r: dicts.get("DI_LEI", {}).get(
                code5._norm_code(r.get("DI_LEI", ""), 6), "")
            or code5._display_di_lei(r.get("DI_LEI", "")),
            "按地类分",
        ),
        _category_line(
            df, code5,
            lambda r: code5._forest_display(
                r.get("SEN_LIN_LB", ""), r.get("Y_SQDJ", ""), r.get("GJGYL_BHDJ", "")),
            "按森林类别分",
        ),
        _category_line(
            df, code5,
            lambda r: "{}保护林地".format(code5._normalize_protect_grade(r.get("BH_DJ", "")))
            if code5._normalize_protect_grade(r.get("BH_DJ", "")) else "",
            "按林地保护等级分",
        ),
        _category_line(
            df, code5,
            lambda r: "{}林地".format(code5._normalize_ownership(r.get("LD_QS", ""))),
            "按林地权属分",
        ),
    ]

    natural = 0.0
    for _, r in df.iterrows():
        if code5._display_origin(r.get("QI_YUAN", "")) == "天然":
            natural += code5._num(r.get("XBMJ", 0))
    paragraphs.append(
        "项目区涉及天然林{}公顷。".format(_area4(natural)) if natural > 0 else "未涉及天然林。"
    )

    paragraphs.append(_vegetation_line(df, code5, dicts))
    paragraphs.append(_discrepancy_line(df, code5, dicts, county))
    paragraphs.append(_protected_area_line(df, code5, county))
    paragraphs.append("项目区拟使用林地不存在未批先占或擅自改变林地用途、采伐林木的行为。")
    paragraphs.append(
        "项目拟使用林地符合《建设项目使用林地审核审批管理办法》"
        "（国家林业局令第35号）中建设项目使用林地的条件和范围。"
    )
    return "\n".join(paragraphs)


# ============================================================
# 报告2/3 表格数据与报告0 汇总
# ============================================================

# 重点保护区域字段 → 报告中的区域类型
_PROTECTED_AREA_TYPE = {
    "ZRBHQ_MC": "自然保护区",
    "SLGY_MC": "自然公园", "SDGY_MC": "自然公园", "FJMSQ_MC": "自然公园",
    "DZGY_MC": "自然公园", "HYGY_MC": "自然公园", "SMGY_MC": "自然公园", "CYGY_MC": "自然公园",
}


def _own_split(df, code5, name_fn, col):
    """按 name_fn(row) 分组，按权属(国有/集体)拆分 col 列合计。
    返回 {name: [total, state, collective]}（仅含 name 非空项）。"""
    groups = {}
    for _, r in df.iterrows():
        nm = name_fn(r)
        if not nm:
            continue
        a = code5._num(r.get(col, 0))
        g = groups.setdefault(nm, [0.0, 0.0, 0.0])
        g[0] += a
        own = code5._normalize_ownership(r.get("LD_QS", ""))
        if own == "国有":
            g[1] += a
        elif own == "集体":
            g[2] += a
    return groups


def _ordered_items(groups, order):
    """按 order 排序，未列入者按合计降序追加。返回 [(name,total,state,collective),...]。"""
    items = [(nm, g[0], g[1], g[2]) for nm, g in groups.items()]
    in_order = [(nm, g[0], g[1], g[2]) for nm, g in groups.items() if nm in order]
    in_order.sort(key=lambda x: order.index(x[0]))
    rest = [(nm, g[0], g[1], g[2]) for nm, g in groups.items() if nm not in order]
    rest.sort(key=lambda x: -x[1])
    return in_order + rest


def _md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "| " + " | ".join("---" for _ in headers) + " |"]
    for r in rows:
        out.append("| " + " | ".join(str(c) for c in r) + " |")
    return "\n".join(out)


def _read_fee_yuan(code5, cname):
    """读取县 可研附表 Fee 表 K6（万元）→ 取整元。失败返回 None。"""
    xlsx = Path(OUTPUT_BASE) / cname / "可研数据" / "可研附表模板.xlsx"
    if not xlsx.exists():
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx), data_only=True)
        ws = wb["项目拟使用林地应缴纳森林植被恢复费测算统计表"]
        val = ws.cell(6, 11).value  # K6 = 总费用（万元）
        if val is None:
            return None
        return int(round(float(val) * 10000))
    except Exception:
        return None


def _eco_region_line(df, code5, county):
    """重点生态区域：按 区域类型×级别 汇总面积。级别从保护区名称含“国家级/省级”判定。"""
    groups = {}  # (类型, 级别) -> area
    for _, r in df.iterrows():
        for f, typ in _PROTECTED_AREA_TYPE.items():
            if f not in df.columns:
                continue
            nm = code5._text(r.get(f, ""))
            if not nm:
                continue
            level = "国家级" if "国家级" in nm else ("省级" if "省级" in nm else "其他")
            groups[(typ, level)] = groups.get((typ, level), 0.0) + code5._num(r.get("XBMJ", 0))
            break
    if not groups:
        return "未涉及国家公园、自然保护区、自然公园等重点生态区域林地。"
    parts = []
    for (typ, level), a in sorted(groups.items()):
        parts.append("{}{}{}公顷".format(typ, level, _area4(a)))
    return "涉及重点生态区域林地：" + "、".join(parts) + "。"


def _di_lei_name(code5, dicts, r):
    return (dicts.get("DI_LEI", {}).get(code5._norm_code(r.get("DI_LEI", ""), 6), "")
            or code5._display_di_lei(r.get("DI_LEI", "")))


def _di_lei_group_rows(df, code5):
    """按规范化地类归为 有林地/国家特别规定灌木林地/其他林地 三组，
    返回 {组名: [合计, 国有, 集体]}。
    有林地=乔木林地+竹林地；其他林地=除有林地与国家特别规定灌木林地之外的林地。"""
    g = {"有林地": [0.0, 0.0, 0.0],
         "国家特别规定灌木林地": [0.0, 0.0, 0.0],
         "其他林地": [0.0, 0.0, 0.0]}
    for _, r in df.iterrows():
        norm = code5._normalize_di_lei(r.get("DI_LEI", ""))
        a = code5._num(r.get("XBMJ", 0))
        own = code5._normalize_ownership(r.get("LD_QS", ""))
        if norm in ("乔木林地", "竹林地"):
            key = "有林地"
        elif norm == "特殊灌林地":
            key = "国家特别规定灌木林地"
        else:
            key = "其他林地"
        g[key][0] += a
        if own == "国有":
            g[key][1] += a
        elif own == "集体":
            g[key][2] += a
    return g


def _forest_class_name(code5, r):
    leaf = code5._forest_leaf(r.get("SEN_LIN_LB", ""), r.get("Y_SQDJ", ""), r.get("GJGYL_BHDJ", ""))
    row = code5._b5_forest_row(leaf)
    return row


def _protect_grade_name(code5, r):
    g = code5._normalize_protect_grade(r.get("BH_DJ", ""))
    return "{}保护林地".format(g) if g else ""


def task_county_tables(cname, df, code5, town_names, vill_names):
    """报告2/3 的表格数据：按县输出用地/林地面积、国有集体、林地类型×权属面积与蓄积、
    按地类/森林类别/保护等级×权属、重点生态区域、天然林、应缴森林植被恢复费。"""
    dicts = _load_dicts(code5)
    first_xian = df.iloc[0].get("XIAN", "") if len(df) else ""
    county = county_name(code5, first_xian, cname)

    total_area = sum(code5._num(v) for v in df.get("XBMJ", []))
    yongdi = _read_xmhx_total_area(code5, cname)
    yongdi_txt = _area4(yongdi) if yongdi is not None else "【用地面积待填】"
    fee = _read_fee_yuan(code5, cname)
    fee_txt = str(fee) if fee is not None else "【费用待填】"

    state_area = collective_area = natural = 0.0
    for _, r in df.iterrows():
        a = code5._num(r.get("XBMJ", 0))
        own = code5._normalize_ownership(r.get("LD_QS", ""))
        if own == "国有":
            state_area += a
        elif own == "集体":
            collective_area += a
        if code5._display_origin(r.get("QI_YUAN", "")) == "天然":
            natural += a

    total_vol = int(round(sum(code5._num(v) for v in df.get("XIAO_BAN_X", []))))

    lines = []
    lines.append("【基本量】{}：用地面积{}公顷；林地面积{}公顷；国有{}公顷；集体{}公顷；"
                 "天然林{}公顷；小班{}个；蓄积{}立方米；应缴森林植被恢复费{}元。".format(
                     county, yongdi_txt, _area4(total_area), _area4(state_area),
                     _area4(collective_area), _area4(natural), len(df), total_vol, fee_txt))

    # 分林地类型、权属 面积表
    g = _own_split(df, code5, lambda r: _b1_land_use(code5, dicts, r), "XBMJ")
    rows = [[nm, _area4(t), _area4(s), _area4(c)] for nm, t, s, c in _ordered_items(g, _B1_LAND_USE_ORDER)]
    rows.append(["合计", _area4(total_area), _area4(state_area), _area4(collective_area)])
    lines.append("\n**分林地类型、权属 面积（公顷）**")
    lines.append(_md_table(["林地类型", "合计", "国有", "集体"], rows))

    # 分林地类型、权属 蓄积表
    gv = _own_split(df, code5, lambda r: _b1_land_use(code5, dicts, r), "XIAO_BAN_X")
    sv = sum(s for _, _, s, _ in [(nm, g[0], g[1], g[2]) for nm, g in gv.items()])
    cv = sum(c for _, _, _, c in [(nm, g[0], g[1], g[2]) for nm, g in gv.items()])
    rowsv = [[nm, int(round(t)), int(round(s)), int(round(c))]
             for nm, t, s, c in _ordered_items(gv, _B1_LAND_USE_ORDER)]
    rowsv.append(["合计", total_vol, int(round(sv)), int(round(cv))])
    lines.append("\n**分林地类型、权属 蓄积（立方米）**")
    lines.append(_md_table(["林地类型", "合计", "国有", "集体"], rowsv))

    # 按地类×权属
    gd = _own_split(df, code5, lambda r: _di_lei_name(code5, dicts, r), "XBMJ")
    gd_items = sorted(((nm, g[0], g[1], g[2]) for nm, g in gd.items()), key=lambda x: -x[1])
    rowsd = [[nm, _area4(t), _area4(s), _area4(c)] for nm, t, s, c in gd_items]
    dg = _di_lei_group_rows(df, code5)
    rowsd.append(["其中：有林地", _area4(dg["有林地"][0]), _area4(dg["有林地"][1]), _area4(dg["有林地"][2])])
    rowsd.append(["其他林地", _area4(dg["其他林地"][0]), _area4(dg["其他林地"][1]), _area4(dg["其他林地"][2])])
    rowsd.append(["合计", _area4(total_area), _area4(state_area), _area4(collective_area)])
    lines.append("\n**按地类分 面积（公顷）**  "
                 "注：有林地=乔木林地+竹林地；其他林地=除有林地与国家特别规定灌木林地之外的林地")
    lines.append(_md_table(["地类", "合计", "国有", "集体"], rowsd))

    # 按森林类别×权属（公益林/商品林）
    gc = _own_split(df, code5, lambda r: _forest_class_name(code5, r), "XBMJ")
    fc_order = ["国家级公益林地", "省级公益林地", "其他公益林地", "重点商品林地", "一般商品林地"]
    rowsc = [[nm, _area4(t), _area4(s), _area4(c)] for nm, t, s, c in _ordered_items(gc, fc_order)]
    rowsc.append(["合计", _area4(total_area), _area4(state_area), _area4(collective_area)])
    lines.append("\n**按森林类别分 面积（公顷）**")
    lines.append(_md_table(["森林类别", "合计", "国有", "集体"], rowsc))

    # 按保护等级×权属
    gp = _own_split(df, code5, lambda r: _protect_grade_name(code5, r), "XBMJ")
    pg_order = ["Ⅰ级保护林地", "Ⅱ级保护林地", "Ⅲ级保护林地", "Ⅳ级保护林地", "Ⅴ级保护林地"]
    rowsp = [[nm, _area4(t), _area4(s), _area4(c)] for nm, t, s, c in _ordered_items(gp, pg_order)]
    rowsp.append(["合计", _area4(total_area), _area4(state_area), _area4(collective_area)])
    lines.append("\n**按林地保护等级分 面积（公顷）**")
    lines.append(_md_table(["保护等级", "合计", "国有", "集体"], rowsp))

    lines.append("\n**重点生态区域林地**")
    lines.append(_eco_region_line(df, code5, county))

    return "\n".join(lines)


def task_report0_summary(county_data, code5, town_names, vill_names):
    """报告0：8县森林植被恢复费明细 + 合计 + 行政许可申请书总量。"""
    rows = []
    grand_area = 0.0
    grand_yongdi = 0.0
    grand_fee = 0
    grand_state = 0.0
    grand_collective = 0.0
    for cname, df in county_data:
        area = sum(code5._num(v) for v in df.get("XBMJ", []))
        yongdi = _read_xmhx_total_area(code5, cname)
        fee = _read_fee_yuan(code5, cname)
        state = collective = 0.0
        for _, r in df.iterrows():
            a = code5._num(r.get("XBMJ", 0))
            own = code5._normalize_ownership(r.get("LD_QS", ""))
            if own == "国有":
                state += a
            elif own == "集体":
                collective += a
        rows.append([cname, _area4(area), fee if fee is not None else ""])
        grand_area += area
        grand_yongdi += (yongdi or 0.0)
        grand_fee += (fee or 0)
        grand_state += state
        grand_collective += collective
    rows.append(["合计", _area4(grand_area), grand_fee])

    lines = []
    lines.append("**森林植被恢复费明细（面积：公顷；金额：元）**")
    lines.append(_md_table(["县（市、区）", "林地面积", "应缴森林植被恢复费"], rows))
    lines.append("")
    lines.append("**行政许可申请书 总量**")
    lines.append("- 项目拟使用土地面积：{}公顷".format(_area4(grand_yongdi)))
    lines.append("- 其中林地面积：{}公顷".format(_area4(grand_area)))
    lines.append("- 国有林地：{}公顷".format(_area4(grand_state)))
    lines.append("- 集体林地：{}公顷".format(_area4(grand_collective)))
    return "\n".join(lines)


TASKS = {
    "forestland_application": {
        "report": "报告1",
        "title": "使用林地申请说明",
        "func": task_forestland_application,
    },
    "used_forestland_unit_detail": {
        "report": "报告2",
        "title": "被使用林地单位明细表",
        "func": task_used_forestland_unit_detail,
    },
    "natural_forest_area": {
        "report": "报告2",
        "title": "天然林林地面积",
        "func": task_natural_forest_area,
    },
    "county_tables": {
        "report": "报告2/3",
        "title": "表格数据（面积/蓄积/地类/类别/保护等级/重点生态区域）",
        "func": task_county_tables,
    },
    "forestland_overview": {
        "report": "报告4",
        "title": "使用林地情况说明",
        "func": task_forestland_overview,
    },
    "review_opinion": {
        "report": "报告4",
        "title": "审查意见",
        "func": task_review_opinion,
    },
    "report0_summary": {
        "report": "报告0",
        "title": "8县汇总（森林植被恢复费 + 申请书总量）",
        "func": task_report0_summary,
        "aggregate": True,
    },
}

# 县内任务渲染顺序：报告1 → 报告2 → 报告2/3 → 报告4；报告0为跨县汇总，单独置顶
PER_COUNTY_ORDER = [
    "forestland_application",
    "used_forestland_unit_detail",
    "natural_forest_area",
    "county_tables",
    "forestland_overview",
    "review_opinion",
]


def parse_args():
    parser = argparse.ArgumentParser(description="按县执行报告小任务并输出 Markdown")
    parser.add_argument(
        "--task",
        default="all",
        choices=["all"] + sorted(TASKS.keys()),
        help="要执行的小任务模块；all 表示全部任务",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help="输出 Markdown 文件名或完整路径",
    )
    return parser.parse_args()


def output_path(name):
    path = Path(name)
    if path.is_absolute():
        return path
    return Path(__file__).with_name(name)


def main():
    args = parse_args()
    code5 = load_code5()
    town_names = code5.parse_township_names(STANDARD_FILE)
    vill_names = code5.parse_village_names(STANDARD_FILE)
    counties = code5.discover_counties(OUTPUT_BASE)

    if not counties:
        print("未发现可处理县目录: {}".format(OUTPUT_BASE))
        return

    county_data = []
    for cname in counties:
        zzy = Path(OUTPUT_BASE) / cname / "林地图斑" / "ZZY.shp"
        df = code5.read_zzy_shp(str(zzy))
        if df is None or len(df) == 0:
            print("跳过 {}: 无 ZZY 数据".format(cname))
            continue
        county_data.append((cname, df))

    out = ["# 报告小任务结果", ""]

    if args.task == "all":
        # 报告0 跨县汇总置顶
        task = TASKS["report0_summary"]
        out.extend(["## {} {}".format(task["report"], task["title"]), "",
                    task["func"](county_data, code5, town_names, vill_names), ""])
        # 每县：报告1 → 报告2 → 报告2/3 → 报告4
        task_keys = PER_COUNTY_ORDER
    else:
        task_keys = [args.task]

    for cname, df in county_data:
        out.append("## {}".format(cname))
        for key in task_keys:
            task = TASKS[key]
            out.extend(["", "### {} {}".format(task["report"], task["title"]), ""])
            if task.get("aggregate"):
                out.extend([task["func"](county_data, code5, town_names, vill_names), ""])
            else:
                out.extend([task["func"](cname, df, code5, town_names, vill_names), ""])

    md_path = output_path(args.output)
    md_path.write_text("\n".join(out).rstrip() + "\n", encoding="utf-8")
    print("\n".join(out))
    print("已生成: {}".format(md_path))


if __name__ == "__main__":
    main()
