import openpyxl
import sys

# Read the xlsx file
wb = openpyxl.load_workbook('ZYY字段名与属性.xlsx', data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
print()

# Print all rows
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    vals = [cell.value for cell in row]
    print(vals)
